"""
DeepEval 评估 - Excel 导出质量
"""
import os
import concurrent.futures
from deepeval import evaluate
from deepeval.test_case import LLMTestCase, SingleTurnParams
from deepeval.metrics import GEval
from openpyxl import load_workbook

# 路径修复：确保 evaluation/ 目录在 sys.path 中
import sys
_eval_dir = os.path.dirname(os.path.abspath(__file__))
if _eval_dir not in sys.path:
    sys.path.insert(0, _eval_dir)

from common import BaseEvaluator, JudgeLLM
from src.agents.api_agent import APITestAgent
from src.tools.rule_manager import get_all_projects, get_module_names


def read_excel_content(filepath):
    """读取 Excel 内容为文本"""
    if not filepath or not os.path.exists(filepath):
        return None
    wb = load_workbook(filepath)
    ws = wb.active
    content = []
    for row in ws.iter_rows(values_only=True):
        row_str = " | ".join([str(cell) if cell else "" for cell in row[:10]])
        if row_str.strip():
            content.append(row_str)
    return "\n".join(content[:30])


def get_test_modules(max_modules: int = None):
    """
    获取所有项目的所有模块。

    参数:
        max_modules: 限制最大模块数，None 表示全部
    """
    projects = get_all_projects()
    modules = []
    for project in projects:
        for module in get_module_names(project):
            modules.append({"project": project, "module": module})
    if max_modules and len(modules) > max_modules:
        modules = modules[:max_modules]
    return modules


class ExcelEvaluator(BaseEvaluator):
    eval_name = "Excel 导出质量评估"
    eval_description = "使用 DeepEval + Qwen-Max 评估 Excel 导出质量"
    eval_dimensions = ["表头规范性", "数据完整性", "格式美观性"]

    @property
    def TEST_SUITE(self):
        return get_test_modules()  # 全部模块，不限量

    def run(self):
        self.print_header()

        agent = APITestAgent()
        test_cases = []
        test_info = []

        all_modules = self.TEST_SUITE
        self.logger.info(f"📊 共 {len(all_modules)} 个模块，开始逐个生成...")

        for idx, test in enumerate(all_modules, 1):
            project = test["project"]
            module = test["module"]
            self.logger.info(f"\n📝 [{idx}/{len(all_modules)}] 生成: {project}/{module}")

            try:
                cases = agent.generate(project, module, "")
            except Exception as e:
                self.logger.info(f"   ❌ 生成失败: {e}")
                continue

            if not cases:
                self.logger.info(f"   ❌ 无用例")
                continue

            filepath = agent.export_excel(cases, project, module)
            excel_content = read_excel_content(filepath)

            if not excel_content:
                self.logger.info(f"   ❌ Excel 导出失败")
                continue

            test_case = LLMTestCase(
                input=f"为{project}/{module}模块导出 Excel 文件",
                actual_output=excel_content
            )
            test_cases.append(test_case)
            test_info.append({"project": project, "module": module, "cases": len(cases)})
            self.logger.info(f"   ✅ 用例数: {len(cases)}")

        if not test_cases:
            self.logger.info("❌ 没有可评估的测试用例")
            return

        excel_metric = GEval(
            name="Excel质量",
            criteria="评估 Excel 导出文件的质量",
            evaluation_steps=[
                "检查是否包含用例编号、用例标题、优先级、请求方法、URL、预期结果列",
                "检查数据是否完整、无空值、无乱码",
                "检查表头是否加粗、列宽是否合适"
            ],
            evaluation_params=[SingleTurnParams.INPUT, SingleTurnParams.ACTUAL_OUTPUT],
            model=JudgeLLM.get_instance(),
            threshold=0.7
        )

        self.logger.info("\n⏳ DeepEval 评估中...")
        try:
            with concurrent.futures.ThreadPoolExecutor() as executor:
                future = executor.submit(evaluate, test_cases, [excel_metric])
                eval_results = future.result(timeout=300)
        except concurrent.futures.TimeoutError:
            self.logger.info("❌ DeepEval 评分超时（超过5分钟）")
            return
        except Exception as e:
            self.logger.info(f"❌ DeepEval 评分失败: {e}")
            return

        self.logger.info("\n" + "=" * 80)
        self.logger.info("📈 评估结果")
        self.logger.info("=" * 80)

        scores = []
        if eval_results and eval_results.test_results:
            for i, metric_result in enumerate(eval_results.test_results):
                info = test_info[i] if i < len(test_info) else {"project": "?", "module": "?"}
                self.logger.info(f"\n📊 {info['project']}/{info['module']}")
                self.logger.info(f"   用例数: {info.get('cases', '?')}")
                for metric in metric_result.metrics_data:
                    score = metric.score
                    scores.append(score)
                    self.logger.info(f"   得分: {score}")
                reason_text = metric.reason if metric.reason else "无详细理由"
                self.logger.info(f"   理由: {reason_text[:200]}")

        # ========== 汇总统计 ==========
        self.logger.info("\n" + "=" * 80)
        self.logger.info("📊 汇总统计")
        self.logger.info("=" * 80)
        self.logger.info(f"\n📋 总模块数: {len(all_modules)}")
        self.logger.info(f"✅ 评估成功: {len(test_cases)}")
        self.logger.info(f"❌ 跳过/失败: {len(all_modules) - len(test_cases)}")
        self.logger.info(f"📊 总用例数: {sum(info.get('cases', 0) for info in test_info)}")

        if scores:
            avg_score = sum(scores) / len(scores)
            self.logger.info(f"\n🏆 综合评分:")
            self.logger.info(f"   平均得分: {avg_score:.2f}")
            self.logger.info(f"   最低得分: {min(scores):.2f}")
            self.logger.info(f"   最高得分: {max(scores):.2f}")
            self.logger.info(f"   ≥ 0.7 阈值通过率: {sum(1 for s in scores if s >= 0.7)}/{len(scores)} ({sum(1 for s in scores if s >= 0.7) / len(scores) * 100:.1f}%)")


def main():
    evaluator = ExcelEvaluator()
    evaluator.run()


if __name__ == "__main__":
    main()
