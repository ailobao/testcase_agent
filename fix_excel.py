# fix_excel.py
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ======================
# 配置
# ======================
DATA_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")


def fix_excel_format(filepath):
    """美化单个 Excel 文件（支持动态列）"""
    if not os.path.exists(filepath):
        print(f"❌ 文件不存在: {filepath}")
        return False

    wb = load_workbook(filepath)
    ws = wb.active
    header_row = 1

    # 获取列名
    col_names = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        col_names.append(val)

    # 默认列宽映射
    default_widths = {
        # 手工用例字段
        "用例ID": 10,
        "标题": 22,
        "前置条件": 20,
        "测试步骤": 40,
        "预期结果": 15,
        "实际结果": 12,
        "优先级": 8,
        # 接口用例字段
        "方法": 8,
        "URL": 40,
        "请求头": 25,
        "参数": 20,
        "请求体": 30,
        "断言": 30,
        "提取变量": 20,
        "前置步骤": 25,
        "后置清理": 20,
    }

    # 动态参数字段（常见）
    dynamic_fields = ["用户名", "密码", "验证码", "商品ID", "数量", "是否勾选",
                      "关键词", "筛选条件", "排序方式", "收货地址", "支付方式", "优惠券"]

    # 设置列宽
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        col_name = col_names[col - 1] if col - 1 < len(col_names) else ""

        if col_name in default_widths:
            ws.column_dimensions[col_letter].width = default_widths[col_name]
        elif col_name in dynamic_fields:
            ws.column_dimensions[col_letter].width = 12
        else:
            ws.column_dimensions[col_letter].width = 15

    # 设置行高
    ws.row_dimensions[header_row].height = 35
    for row in range(header_row + 1, ws.max_row + 1):
        ws.row_dimensions[row].height = 25

    # 表头样式（蓝底白字居中）
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 数据样式
    data_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 数据对齐：短字段居中，长字段左对齐
    center_cols = ["用例ID", "验证码", "实际结果", "优先级", "方法", "状态码"]

    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border

            col_name = col_names[col - 1] if col - 1 < len(col_names) else ""
            if col_name in center_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb.save(filepath)
    print(f"✅ 美化完成: {filepath}")
    return True


def fix_all_excel_in_folder(folder_path):
    """批量美化文件夹内所有 Excel"""
    if not os.path.exists(folder_path):
        print(f"❌ 文件夹不存在: {folder_path}")
        return

    fixed = []
    for f in os.listdir(folder_path):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            filepath = os.path.join(folder_path, f)
            print(f"处理: {f}")
            if fix_excel_format(filepath):
                fixed.append(f)

    print(f"\n✅ 批量美化完成！共处理 {len(fixed)} 个文件")


if __name__ == "__main__":
    print("=" * 50)
    print("Excel 美化工具")
    print("=" * 50)
    print(f"目标文件夹: {DATA_FOLDER}")
    print("=" * 50)

    fix_all_excel_in_folder(DATA_FOLDER)

    input("\n按 Enter 键退出...")